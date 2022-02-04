import argparse
import os
import hashlib
import openpyxl
import xlsxwriter
import requests
from OTXv2 import OTXv2
import IndicatorTypes
import configparser
def walk_directory(args,worksheet):
    #Common file extensions that are not asociated with malwares.
    common_files=['.evtx','.pcapng','.pcap','.txt','.log','.sql','.jpg','.jpeg','.png','.xml','.mp3','.mp4','.db','.css','.lnk','.pf','.mov','.avi','.json','.yml']
    if os.path.isdir(args.input_path):
        row = 1
        for root,dirs,files in os.walk(args.input_path):

            for file_name in files:
                file=os.path.join(root,file_name)
                file_split = os.path.splitext(file)
                if file_split[1] not in common_files:
                 with open(file, 'rb') as target_file:
                        # read complete file as bytes
                        bytes = target_file.read()
                        hashtype=args.algo
                        if( hashtype =='md5' or hashtype=='MD5'):

                            filehash = hashlib.md5(bytes).hexdigest()
                        elif ( hashtype =='SHA1' or hashtype=='sha1'):
                           filehash = hashlib.sha1(bytes).hexdigest()
                        elif ( hashtype =='SHA256' or hashtype=='sha256'):
                         filehash = hashlib.sha256(bytes).hexdigest()
                        elif (hashtype==None):
                            filehash = hashlib.md5(bytes).hexdigest()
                        else:
                            print("\n")
                            print("Wrong Hashing Algorithm selected.Possible hashing alogs are [md5,sha1,sha256].\n")
                            print(" Example: udham.py  -d E:\malwares -m auto -o E:\majid.xlsx --algo sha1")
                            exit()
                        data=[file,filehash]
                        write_data(data,worksheet,row)
                        row+=1
    else:
        print("Directory provided does not exist. Kindly provide correct Path.")
        print(Example)
        exit(0)

def write_file(args):
    c_col=0
    hashtype=''
    if (args.algo==None):
         hashtype='Hash'
    columns=['File_Path',hashtype,'VT_Malicious_Count','VT_Suspicious_Count','VT_Harmless_Count','OTX Pulses Count']
    check_extension= os.path.splitext(args.output_file)
    if '.xlsx' in check_extension[1]:
        workbook = xlsxwriter.Workbook(args.output_file, {'constant_memory': True})
        bold = workbook.add_format({'bold': True})
        worksheet = workbook.add_worksheet()
        for headings in columns:
           worksheet.write(0,c_col,headings,bold)
           c_col=c_col+1
    else:
        print("\n")
        print("Provide correct Path and Name for Output file. Only XLS file type is allowed.")
        print(Example)
        exit(0)
    if (args.input_path!=None):
        walk_directory(args, worksheet)

    workbook.close()

def write_data(data, worksheet, row):
    col=0
    for items in data:
        worksheet.write(row, col, items)
        col=col+1
def get_OTXverdict(args,hash):
    if (args.algo=='md5') or (args.algo=='MD5'):
        hash_type = IndicatorTypes.FILE_HASH_MD5
    elif (args.algo=='sha1') or (args.algo=='SHA1'):
        hash_type = IndicatorTypes.FILE_HASH_SHA1
    elif (args.algo=='sha256') or (args.algo=='SHA256'):
        hash_type = IndicatorTypes.FILE_HASH_SHA256
    elif (args.algo==None):
        hash_type = IndicatorTypes.FILE_HASH_SHA1
    try:
        result = otx.get_indicator_details_by_section(hash_type, hash, 'general')
        count = result['pulse_info']['count']
        return count
    except Exception as e:
        return str(e)

def get_vtverdict(hash):
    try:
        response=requests.request("GET",vt_url+hash,headers=(vt_headers))
        converted = response.json()
    except Exception as e:
        print(e)
    try:
        malicious = converted["data"][0]['attributes']['last_analysis_stats']['malicious']
        suspicious = converted["data"][0]['attributes']['last_analysis_stats']['suspicious']
        harmless   = converted["data"][0]['attributes']['last_analysis_stats']['harmless']
        undected   = converted["data"][0]['attributes']['last_analysis_stats']['undetected']
        results=[malicious,suspicious,harmless,undected]
        return results
    except:
        return ['Not Found','Not Found','Not Found']

def check_args(arg):

    if arg.mode ==('save' ) or arg.mode ==('auto'):
        if  not (os.path.isdir(arg.input_path)):
            print("\n")
            print("Provided Directory Not Found. \n")
            print(parser.print_help())
            exit()
    elif arg.mode =='scan':
       try:
        if not (os.path.isfile(arg.input_file)):
            print("\n")
            print("Provided File does not exist.\n")
            print(parser.print_help())
            exit()
       except:
           print("\n")
           print("With Scan mode you can only scan files not Directories. To scan a directory use auto mode. \n")
           print( 'Usage : udham.py -f E:\\file.txt -m scan')

    else:
        print("\n")
        print("Wrong Mode Selected. \n")
        print(parser.print_help())
        exit()
def get_xlsxdata(args):
    outfile=openpyxl.load_workbook(args.output_file)
    #Create object of active sheet
    sheet=outfile.active
    #total rows in active sheet
    total_rows=sheet.max_row
    for row in range(total_rows):
        if (sheet.cell(row + 2, 2).value)!=None:
            results=get_vtverdict(sheet.cell(row+2,2).value)
            otx_result=get_OTXverdict(args,sheet.cell(row+2,2).value)

            sheet.cell(row+2,3).value=results[0]
            sheet.cell(row+2,4).value=results[1]
            sheet.cell(row+2,5).value=results[2]
            sheet.cell(row+2,6).value=otx_result
    outfile.save(args.output_file)
    outfile.close()
def getmodes(args):
    if args.mode=='save' :
        write_file(args)

    elif args.mode=='auto':
        write_file(args)
        get_xlsxdata(args)
    elif args.mode=='scan':
        if (os.path.splitext(args.input_file))[1]=='.txt':
            txt_readfile(args)
        elif(os.path.splitext(args.input_file))[1]=='.xlsx':
            xlsx_readfile(args)


def txt_readfile(args):
    row=2
    if (os.path.isfile(args.input_file)):
        write_file(args)
        outfile = openpyxl.load_workbook(args.output_file)
        sheet = outfile.active
        hashfile = open(args.input_file, 'r')
        for hashline in hashfile.readlines():
            result = get_vtverdict(hashline.rstrip())
            otx_result = get_OTXverdict(args, hashline.rstrip())
            sheet.cell(row, 1).value = args.input_file
            sheet.cell(row, 2).value = hashline
            sheet.cell(row, 3).value = result[0]
            sheet.cell(row, 4).value = result[1]
            sheet.cell(row, 5).value = result[2]
            sheet.cell(row, 6).value = otx_result
            row+=1


    outfile.save(args.output_file)
    outfile.close()
def xlsx_readfile(args):
    if (os.path.isfile(args.input_file)):
        write_file(args)
        outfile = openpyxl.load_workbook(args.output_file)
        sheet = outfile.active
        infile = openpyxl.load_workbook(args.input_file)
        insheet = infile.active
        max_rows = insheet.max_row
        for row in range(max_rows):
            filehash = insheet.cell(row+2, 2).value
            if hash != None:
                result = get_vtverdict(filehash)
                otx_result = get_OTXverdict(args, filehash)
                sheet.cell(row+1, 1).value = args.input_file
                sheet.cell(row+2, 2).value = filehash
                sheet.cell(row+2, 3).value = result[0]
                sheet.cell(row+2, 4).value = result[1]
                sheet.cell(row+2, 5).value = result[2]
                sheet.cell(row+2, 6).value = otx_result

    outfile.save(args.output_file)
    outfile.close()

def get_args():

    #Command Line Arguments
    parser.add_argument('-f','--file', dest='input_file',help=
    ' Provide Exact Path for Text or CSV file to read hashes from.')
    parser.add_argument('-d',dest='input_path',help=
    'Provide Directory that you want to traverse for hash calculation.')
    parser.add_argument('--algo',dest='algo',help='Algorithm to calculate hash possible algos [md5,sha1,sha256]')
    parser.add_argument('-o','--output',dest='output_file',help=
    'OutPut file Path. Make sure you have permission to write.')
    parser.add_argument('-m','--mode',dest='mode',help='Possible mode are [save,scan,auto]')
    arguments = parser.parse_args()
    return arguments

if __name__ == '__main__':
    # Author : Majid Jahangeer
    #Date : 05-02-2022
    #Version 1.0
    
    import ast
    config = configparser.ConfigParser()
    config.read('config.ini')
    parser = argparse.ArgumentParser()
    # OTX URL Address
    OTX_SERVER = 'https://otx.alienvault.com/'
    # Your OTX API ket. By default you can search for 10K requests per hour.
    OTX_API_KEY = config['OTX']['OTX_API_KEY']
    otx = OTXv2(OTX_API_KEY, server=OTX_SERVER)
    # Virus total url and header info.Give your VT API key.
    vt_url = "https://www.virustotal.com/api/v3/search?query="
    vt_headers = config['VirusTotal']['vt_headers']
    vt_headers=ast.literal_eval(vt_headers)
    if not (vt_headers['x-apikey']) and OTX_API_KEY:
        print("\n")
        print("Please provide Virus Total and OTX API KEY in config.ini file.Config.ini must be in the same directory as odhum.py.")
        print(parser.print_help())
        exit()
    Example= 'Usage : python udham.py -d E:\malwares -o E:\output\outfile.xlsx'
    args=get_args()
    check_args(args)
    getmodes(args)
